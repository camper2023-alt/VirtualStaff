import PySimpleGUI as sg
import openpyxl
import glob

#関数の定義
#ボタンを押すと、大会名＝シート名のシートを作成する。
def player_register(v,win):
    wb = openpyxl.load_workbook('KP_Sheets.xlsx')
    wb.create_sheet('{}'.format(v["Sheet_name"]))
    sheet2 = wb.get_sheet_by_name('{}'.format(v["Sheet_name"]))
    sheet2.cell(row=1, column=10).value = 0
    wb.save('KP_Sheets.xlsx')
    win["exp0"].Update('{}用シートを作成しました。'.format(v["Sheet_name"]))


#ボタンを押すと、新しいファイルを作成する。
def create_file():
    li_file = glob.glob("*.xlsx")
    if 'KP_Sheets.xlsx' in li_file:
        wb = openpyxl.load_workbook('KP_Sheets.xlsx')
        sheet0 = wb.get_sheet_by_name('Players_list')
        li_player = [pl.value for pl in list(sheet0.columns)[0]]
        li_sheet = wb.sheetnames
    else:
        wb = openpyxl.Workbook()
        sheet1 = wb.active
        sheet1.title = 'Players_list'
        li_sheet = wb.sheetnames
        wb.save('KP_Sheets.xlsx')


#ボタンを押すと、入力したパーティが指定したシートに書き込まれる。
def set_team(v,win):
    wb = openpyxl.load_workbook('KP_Sheets.xlsx')
    sheet3 = wb.get_sheet_by_name('{}'.format(v["Sheet_name"]))
    num = int(sheet3.cell(row=1, column=10).value)
    for i in range(1,7):
        sheet3.cell(row=num+1, column=i).value = v["poke{}".format(i)]
    sheet3.cell(row=num+1, column=7).value = v["rental1"]
    sheet3.cell(row=num+1, column=8).value = v["name"]
    sheet3.cell(row=num+1, column=9).value = num + 1
    sheet3.cell(row=1, column=10).value = num + 1
    wb.save('KP_Sheets.xlsx')
    win["exp0"].Update('{}さんのパーティが登録されました。'.format(v["name"]))
    for i in range(1,7):
        win["poke{}".format(i)].Update('')
    win["name"].Update('')
    win["rental1"].Update('')
    
    
    


#ボタンを押すと、入力済みのパーティのデータを呼び出す。
def open_team(v,win):
    Flag = False
    wb =openpyxl.load_workbook('KP_Sheets.xlsx')
    sheet4 = wb.get_sheet_by_name('{}'.format(v["Sheet_name"]))
    num = int(sheet4.cell(row=1, column=10).value)
    for i in range(1,num+1):
        na = sheet4.cell(row=i, column=8).value
        if na == v["name"]:
            num = i
            Flag = True
            break
    if Flag:
        for i in range(1,7):
            win["poke{}".format(i)].Update(sheet4.cell(row=num, column=i).value)
    else:
        win["exp0"].Update('{}さんのパーティは登録されていません。'.format(v["name"]))
        

#ボタンを押すと、入力済みのパーティのデータからKPを集計し、表示する。
def sum_KP(v,win):
    wb =openpyxl.load_workbook('KP_Sheets.xlsx')
    sheet5 = wb.get_sheet_by_name('{}'.format(v["Sheet_name"]))
    num = int(sheet5.cell(row=1, column=10).value)
    dic_KP = {}
    for i in range(1,num+2):
        for j in range(1,7):
            if sheet5.cell(row=i, column=j).value not in dic_KP.keys():
                dic_KP[sheet5.cell(row=i, column=j).value] = 1
            else:
                dic_KP[sheet5.cell(row=i, column=j).value] += 1
    del dic_KP[None]

    co_KP = []
    for i in dic_KP:
        if dic_KP[i] not in co_KP:
            co_KP.append(dic_KP[i])
    
    co_KP.sort(reverse=True)
    f = open('{}KPまとめ.txt'.format(v["Sheet_name"]), 'a')
    for i in range(len(co_KP)):
        li_KP = [key for key, value in dic_KP.items() if value == co_KP[i]]
        f.write('{}:'.format(co_KP[i]))
        for j in range(len(li_KP)):
            if j+1 == len(li_KP):
                f.write('{}\n'.format(li_KP[j]))
            else:
                f.write('{},'.format(li_KP[j]))
    f.close()
    win["exp0"].Update('KPの集計ができました。')


#ボタンを押すと、入力欄を空にする。
def clear(win):
    for i in range(1,7):
        win["poke{}".format(i)].Update('')
    win["name"].Update('')
    win["rental1"].Update('')


#ボタンを押すと、既に入力されたパーティのデータを削除する。
def del_team(v,win):
    wb =openpyxl.load_workbook('KP_Sheets.xlsx')
    sheet6 = wb.get_sheet_by_name('{}'.format(v["Sheet_name"]))
    num = int(sheet6.cell(row=1, column=10).value)
    Flag = False
    for i in range(1,num+1):
        na = sheet6.cell(row=i, column=8).value
        if na == v["name"]:
            num = i
            Flag = True
            break
    if Flag:
        for i in range(1,10):
            sheet6.cell(row=num, column=i).value = None
            win["exp0"].Update('{}さんのパーティが削除されました。'.format(v["name"]))
    else:
        win["exp0"].Update('{}さんのパーティは登録されていません。'.format(v["name"]))
    wb.save('KP_Sheets.xlsx')

#ボタンを押すと、パーティ未提出の人数を表示する。ある程度少なくなったら名前を表示する。
def check_set_numbers(v,win,li_all_players):
    wb = openpyxl.load_workbook('KP_Sheets.xlsx')
    sheet7 = wb.get_sheet_by_name('{}'.format(v["Sheet_name"]))
    num = int(sheet7.cell(row=1, column=10).value)
    for i in range(1,num+2):
        na = sheet7.cell(row=i, column=8).value
        if na in li_all_players:
            li_all_players.remove(sheet7.cell(row=i, column=8).value)
    if len(li_all_players) >= 10:
        win["exp0"].Update('未提出者は残り{}人です。'.format(len(li_all_players)))
    elif len(li_all_players) == 0:
        win["exp0"].Update('参加者全員の提出が完了しました。')
    else:
        win["exp0"].Update('未提出者は以下の通りです。{}'.format(li_all_players))