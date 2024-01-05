import PySimpleGUI as sg
import openpyxl
import glob
import pokemon_database as pd
import function_KP



#必要なリスト、辞書の定義
#参加者のリスト
li_player = []
li_all_players = []

#シートのリスト
li_sheet = []

#元のエクセルファイル呼び出し、なければ新規作成
li_file = glob.glob("*.xlsx")
if 'KP_Sheets.xlsx' in li_file:
    wb = openpyxl.load_workbook('KP_Sheets.xlsx')
    sheet0 = wb.get_sheet_by_name('Players_list')
    li_player = [pl.value for pl in list(sheet0.columns)[0]]
    li_all_players = [pl.value for pl in list(sheet0.columns)[0]]
    li_sheet = wb.sheetnames
    li_sheet.remove('Players_list')
else:
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = 'Players_list'
    wb.save('KP_Sheets.xlsx')
    li_file = glob.glob("*.xlsx")


#コンボボックス用のデータベースのリスト
li_poke = sorted(pd.pokemon)


#アプリのレイアウト

layout = [[sg.T("HN", k="hn"), sg.Combo(li_player, size=(16, 8), k="name"), sg.T("オフ名", k="file"), sg.Combo(li_sheet, size=(16, 4), k="Sheet_name")],
          [sg.T("", size=(12,1), k="0"), sg.T("ポケモン１", size=(15,1), k="1", justification="center"), sg.T("ポケモン２", size=(16,1), k="2", justification="center"), sg.T("ポケモン３", size=(15,1), k="3", justification="center"), sg.T("ポケモン４", size=(16,1), k="4", justification="center"), sg.T("ポケモン５", size=(15,1), k="5", justification="center"), sg.T("ポケモン６", size=(16,1), k="6", justification="center")],
          [sg.T("種族名", size=(12,1), k="poke0"), sg.Combo(li_poke, size=(16, 12), k="poke1"), sg.Combo(li_poke, size=(16, 12), k="poke2"), sg.Combo(li_poke, size=(16, 12), k="poke3"), sg.Combo(li_poke, size=(16, 12), k="poke4"), sg.Combo(li_poke, size=(16, 12), k="poke5"), sg.Combo(li_poke, size=(16, 12), k="poke6")],
          [sg.T("レンタルID", size=(12,1), k="rental0"), sg.InputText(size=(16, 6), k="rental1")],
          [sg.T("", size=(12,1), k="btn0"), sg.B("シート作成", size=(15,2), k="btn1"), sg.B("パーティ登録", size=(15,2), k="btn2"), sg.B("パーティ開示", size=(15,2), k="btn3"), sg.B("入力欄クリア", size=(15,2), k="btn4"), sg.B("KP集計", size=(15,2), k="btn5"), sg.B("パーティ削除", size=(15,2), k="btn6")],
          [sg.T("", size=(12,1), k="btn7"), sg.B("提出チェック", size=(15,2), k="btn8")],
          [sg.T("", size=(120,2), k="exp0")],
          [sg.T("{}".format(li_player), size=(120,5), k="exp1")],
          [sg.T("※A=アローラ、G=ガラル、H=ヒスイ、P=パルデア、O=オリジン", size=(120,1), k="exp2")],
          [sg.T("作った人:おでん、対応環境:SV藍の円盤", size=(120,1), k="exp3")]
          ]

win = sg.Window("VirtualStaff1.0", layout)

#機能
while True:
    e,v = win.read()
    if e == "btn1":
        function_KP.player_register(v,win)
    if e == "btn2":
        function_KP.set_team(v,win)
    if e == "btn3":
        function_KP.open_team(v,win)
    if e == "btn4":
        function_KP.clear(win)
    if e == "btn5":
        function_KP.sum_KP(v,win)
    if e == "btn6":
        function_KP.del_team(v,win)
    if e == "btn8":
        function_KP.check_set_numbers(v,win,li_all_players)
    if e == None:
        break
win.close()